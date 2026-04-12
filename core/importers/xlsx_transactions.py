from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, datetime
from pathlib import Path
from typing import Any

from core.domain import TransactionInput, TransferInput
from core.persistence.repositories.accounts_repo import AccountsRepository
from core.persistence.repositories.categories_repo import CategoriesRepository
from core.services.transactions import TransactionsService


@dataclass(frozen=True)
class XLSXImportResult:
    imported_count: int
    deleted_count: int
    import_period_key: str
    year_month_keys: tuple[str, ...]


@dataclass(frozen=True)
class TransferRule:
    name: str
    match_category: str
    match_description: str
    from_account_number: str
    from_account_type: str = "checking"
    to_account_number: str = ""
    to_account_type: str = "savings"
    enabled: bool = True


class XLSXTransactionImporter:
    def __init__(
        self,
        transactions_service: TransactionsService,
        categories_repo: CategoriesRepository,
        accounts_repo: AccountsRepository,
        *,
        transfer_rules: list[dict[str, Any]] | None = None,
        logger=None,
    ) -> None:
        self.transactions_service = transactions_service
        self.categories_repo = categories_repo
        self.accounts_repo = accounts_repo
        self.transfer_rules = self._normalize_transfer_rules(transfer_rules or [])
        self.logger = logger

    @staticmethod
    def _normalize_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    @staticmethod
    def _section_kind(label: str | None, sequence_index: int) -> str:
        text = (label or "").strip().lower()
        if "income" in text:
            return "income"
        if "expense" in text:
            return "expense"
        # Fallback when section labels are missing: first table expense, second income.
        return "expense" if sequence_index == 0 else "income"

    def _find_sections(self, ws) -> list[dict[str, Any]]:
        sections: list[dict[str, Any]] = []
        max_scan_rows = min(ws.max_row, 120)
        max_scan_cols = max(7, ws.max_column)
        header_pattern = ["date", "amount", "description", "category"]
        optional_headers = {
            "account",
            "subscription",
            "tax",
            "type",
            "payment type",
            "payment_type",
            "note",
            "notes",
        }

        for row in range(1, max_scan_rows + 1):
            for col in range(1, max_scan_cols - 2):
                headers = [
                    self._normalize_text(ws.cell(row=row, column=col + idx).value)
                    for idx in range(4)
                ]
                if headers != header_pattern:
                    continue

                column_map: dict[str, int] = {
                    "date": col,
                    "amount": col + 1,
                    "description": col + 2,
                    "category": col + 3,
                }
                next_col = col + 4
                while next_col <= max_scan_cols:
                    header_name = self._normalize_text(ws.cell(row=row, column=next_col).value)
                    if not header_name:
                        break
                    if header_name in optional_headers:
                        if header_name == "notes":
                            canonical_header = "note"
                        elif header_name in {"payment type", "payment_type"}:
                            canonical_header = "type"
                        else:
                            canonical_header = header_name
                        if canonical_header in column_map:
                            next_col += 1
                            continue
                        column_map[canonical_header] = next_col
                        next_col += 1
                        continue
                    break

                section_label = None
                for lookback in (1, 2, 3):
                    r = row - lookback
                    if r < 1:
                        continue
                    candidate = ws.cell(row=r, column=col).value
                    if candidate is not None and str(candidate).strip():
                        section_label = str(candidate).strip()
                        break

                sections.append(
                    {
                        "kind": self._section_kind(section_label, len(sections)),
                        "header_row": row,
                        "start_col": col,
                        "column_map": column_map,
                    }
                )

        if not sections:
            raise ValueError(
                "Could not find expected Transactions worksheet tables. "
                "Expected headers: Date, Amount, Description, Category."
            )
        return sections

    @staticmethod
    def _parse_date(value: Any, row_num: int) -> str:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()

        text = str(value).strip() if value is not None else ""
        if not text:
            raise ValueError(f"Missing date at row {row_num}")

        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        raise ValueError(f"Invalid date '{text}' at row {row_num}")

    @staticmethod
    def _parse_amount_cents(value: Any, row_num: int) -> int:
        if isinstance(value, (int, float)):
            amount = float(value)
        else:
            text = str(value).strip() if value is not None else ""
            if not text:
                raise ValueError(f"Missing amount at row {row_num}")
            is_parenthetical_negative = text.startswith("(") and text.endswith(")")
            cleaned = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
            try:
                amount = float(cleaned)
            except ValueError as exc:
                raise ValueError(f"Invalid amount '{text}' at row {row_num}") from exc
            if is_parenthetical_negative:
                amount *= -1

        cents = int(round(amount * 100))
        if cents == 0:
            raise ValueError(f"Amount may not be 0 at row {row_num}")
        return cents

    @staticmethod
    def _normalize_description(value: Any) -> str:
        text = str(value).strip() if value is not None else ""
        if text in {"...", "…"}:
            return ""
        return text

    @staticmethod
    def _parse_bool(value: Any, default: bool, row_num: int, field_name: str) -> bool:
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0

        text = str(value).strip().lower()
        if text == "":
            return default
        if text in {"true", "t", "yes", "y", "1", "x", "checked", "check", "✓"}:
            return True
        if text in {"false", "f", "no", "n", "0", "unchecked", "off"}:
            return False
        raise ValueError(f"Invalid {field_name} value '{value}' at row {row_num}")

    @staticmethod
    def _normalize_account_type(value: Any, default_account_type: str, row_num: int) -> str:
        if value is None:
            return default_account_type

        text = str(value).strip().lower()
        if not text:
            return default_account_type

        compact = "".join(ch for ch in text if ch.isalnum())
        if compact in {"cash"}:
            return "cash"
        if compact in {"checking", "check"}:
            return "checking"
        if compact in {"credit", "creditcard", "card"}:
            return "credit"
        if compact in {"savings", "saving"}:
            return "savings"
        raise ValueError(
            "Invalid account value "
            f"'{value}' at row {row_num}. Expected cash/checking/credit/savings."
        )

    @staticmethod
    def _normalize_transfer_rules(raw_rules: list[dict[str, Any]]) -> list[TransferRule]:
        rules: list[TransferRule] = []
        for idx, item in enumerate(raw_rules):
            if not isinstance(item, dict):
                continue
            match_category = str(item.get("match_category", "")).strip()
            match_description = str(item.get("match_description", "")).strip()
            from_account_number = str(item.get("from_account_number", "")).strip()
            to_account_number = str(item.get("to_account_number", "")).strip()
            if not match_category or not match_description or not from_account_number or not to_account_number:
                continue
            from_account_type = str(item.get("from_account_type", "checking")).strip().lower() or "checking"
            to_account_type = str(item.get("to_account_type", "savings")).strip().lower() or "savings"
            if from_account_type not in {"cash", "checking", "credit", "savings"}:
                from_account_type = "checking"
            if to_account_type not in {"cash", "checking", "credit", "savings"}:
                to_account_type = "savings"
            name = str(item.get("name", "")).strip() or f"Rule {idx + 1}"
            rules.append(
                TransferRule(
                    name=name,
                    match_category=match_category,
                    match_description=match_description,
                    from_account_number=from_account_number,
                    from_account_type=from_account_type,
                    to_account_number=to_account_number,
                    to_account_type=to_account_type,
                    enabled=bool(item.get("enabled", True)),
                )
            )
        return rules

    @staticmethod
    def _normalize_account_number(value: str | None) -> str:
        return str(value or "").strip().casefold()

    @staticmethod
    def _resolve_account_id_by_number(
        *,
        account_number: str,
        account_type: str,
        account_rows_by_number: dict[str, dict[str, Any]],
    ) -> int | None:
        normalized_number = XLSXTransactionImporter._normalize_account_number(account_number)
        if not normalized_number:
            return None
        row = account_rows_by_number.get(normalized_number)
        if row is None:
            return None
        normalized_type = str(account_type or "").strip().lower()
        row_type = str(row.get("account_type") or "").strip().lower()
        if normalized_type and row_type and row_type != normalized_type:
            return None
        return int(row["account_id"])

    def _match_transfer_rule(self, category_name: str, description: str) -> TransferRule | None:
        category_key = str(category_name or "").strip().casefold()
        description_key = str(description or "").strip().casefold()
        if not category_key or not description_key:
            return None
        for rule in self.transfer_rules:
            if not rule.enabled:
                continue
            if rule.match_category.strip().casefold() != category_key:
                continue
            rule_desc = rule.match_description.strip().casefold()
            if rule_desc and rule_desc in description_key:
                return rule
        return None

    @staticmethod
    def _infer_import_period_key(txn_dates: list[str]) -> str:
        counts: dict[str, int] = {}
        first_seen: dict[str, int] = {}
        for index, txn_date in enumerate(txn_dates):
            key = str(txn_date)[:7]
            counts[key] = counts.get(key, 0) + 1
            if key not in first_seen:
                first_seen[key] = index

        # Pick the dominant year-month; ties go to earliest appearance in the sheet.
        return min(
            counts.keys(),
            key=lambda key: (-counts[key], first_seen[key]),
        )

    def import_file(
        self,
        xlsx_path: Path,
        default_account: str = "Checking",
        replace_monthly_baseline: bool = True,
    ) -> XLSXImportResult:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "openpyxl is required for XLSX imports. Install it with: pip install openpyxl"
            ) from exc

        wb = load_workbook(xlsx_path, data_only=True)
        if "Transactions" not in wb.sheetnames:
            raise ValueError("Worksheet 'Transactions' was not found in workbook.")
        ws = wb["Transactions"]

        account_ids_by_type: dict[str, int] = {}
        account_rows_by_number: dict[str, dict[str, Any]] = {}
        for row in self.accounts_repo.list_active():
            account_type = str(row["account_type"]).strip().lower()
            account_id = int(row["account_id"])
            if account_type not in account_ids_by_type:
                account_ids_by_type[account_type] = account_id
            account_number_key = self._normalize_account_number(str(row.get("account_number") or ""))
            if account_number_key and account_number_key not in account_rows_by_number:
                account_rows_by_number[account_number_key] = dict(row)
        for account_type, account_name in (
            ("cash", "Cash"),
            ("checking", "Checking"),
            ("credit", "Credit Card"),
            ("savings", "Savings"),
        ):
            if account_type not in account_ids_by_type:
                account_ids_by_type[account_type] = self.accounts_repo.upsert(account_name, account_type)

        sections = self._find_sections(ws)
        row_limit = ws.max_row
        parsed_items: list[tuple[str, TransactionInput | TransferInput]] = []
        parsed_dates: list[str] = []

        for section in sections:
            kind = section["kind"]
            row = int(section["header_row"]) + 1
            column_map = dict(section.get("column_map", {}))
            date_col = int(column_map.get("date", section["start_col"]))
            amount_col = int(column_map.get("amount", date_col + 1))
            description_col = int(column_map.get("description", date_col + 2))
            category_col = int(column_map.get("category", date_col + 3))
            account_col = int(column_map["account"]) if "account" in column_map else None
            subscription_col = int(column_map["subscription"]) if "subscription" in column_map else None
            tax_col = int(column_map["tax"]) if "tax" in column_map else None
            payment_type_col = int(column_map["type"]) if "type" in column_map else None
            note_col = int(column_map["note"]) if "note" in column_map else None
            blank_streak = 0

            while row <= row_limit:
                date_value = ws.cell(row=row, column=date_col).value
                amount_value = ws.cell(row=row, column=amount_col).value
                desc_value = ws.cell(row=row, column=description_col).value
                category_value = ws.cell(row=row, column=category_col).value
                account_value = ws.cell(row=row, column=account_col).value if account_col else None
                subscription_value = (
                    ws.cell(row=row, column=subscription_col).value if subscription_col else None
                )
                tax_value = ws.cell(row=row, column=tax_col).value if tax_col else None
                payment_type_value = (
                    ws.cell(row=row, column=payment_type_col).value if payment_type_col else None
                )
                note_value = ws.cell(row=row, column=note_col).value if note_col else None

                if all(
                    v is None or (isinstance(v, str) and not v.strip())
                    for v in (date_value, amount_value, desc_value, category_value)
                ):
                    blank_streak += 1
                    if blank_streak >= 2:
                        break
                    row += 1
                    continue
                blank_streak = 0

                txn_date = self._parse_date(date_value, row)
                amount_cents = self._parse_amount_cents(amount_value, row)
                if kind == "expense" and amount_cents > 0:
                    amount_cents *= -1
                elif kind == "income" and amount_cents < 0:
                    amount_cents = abs(amount_cents)

                description = self._normalize_description(desc_value)
                category_name = (
                    str(category_value).strip()
                    if category_value is not None and str(category_value).strip()
                    else ""
                )
                payment_type = str(payment_type_value).strip() if payment_type_value is not None else ""
                note_text = str(note_value).strip() if note_value is not None else ""
                internal_payee = description or "Transaction"
                default_account_type = "checking" if kind == "income" else "credit"
                account_type = self._normalize_account_type(account_value, default_account_type, row)
                account_id = account_ids_by_type[account_type]
                is_subscription = False
                if kind == "expense":
                    is_subscription = self._parse_bool(
                        subscription_value,
                        default=False,
                        row_num=row,
                        field_name="Subscription",
                    )

                tax_default = True if kind == "income" else False
                tax_flag = self._parse_bool(
                    tax_value,
                    default=tax_default,
                    row_num=row,
                    field_name="Tax",
                )
                tax_category = "Other" if (kind == "expense" and tax_flag) else None

                category = self.categories_repo.find_by_name(category_name) if category_name else None
                category_id = int(category["category_id"]) if category else None

                source_uid = f"Transactions:{kind}:{row}"
                transfer_rule = self._match_transfer_rule(category_name, description) if kind == "expense" else None
                if transfer_rule is not None:
                    from_account_id = self._resolve_account_id_by_number(
                        account_number=transfer_rule.from_account_number,
                        account_type=transfer_rule.from_account_type,
                        account_rows_by_number=account_rows_by_number,
                    )
                    to_account_id = self._resolve_account_id_by_number(
                        account_number=transfer_rule.to_account_number,
                        account_type=transfer_rule.to_account_type,
                        account_rows_by_number=account_rows_by_number,
                    )
                    if from_account_id is not None and to_account_id is not None:
                        if from_account_id == to_account_id:
                            if self.logger is not None:
                                self.logger.warning(
                                    "Transfer rule '%s' matched but resolved to the same account "
                                    "(account_id=%s); transfer skipped and row imported as expense.",
                                    transfer_rule.name,
                                    from_account_id,
                                )
                        else:
                            transfer = TransferInput(
                                txn_date=txn_date,
                                amount_cents=abs(amount_cents),
                                from_account_id=from_account_id,
                                to_account_id=to_account_id,
                                payee=internal_payee,
                                category_id=category_id,
                                description=description or category_name or "Transfer",
                                note=note_text or None,
                                source_system="xlsx_import",
                                source_uid=source_uid,
                                payment_type=None,
                            )
                            parsed_items.append(("transfer", transfer))
                            parsed_dates.append(txn_date)
                            row += 1
                            continue
                    if self.logger is not None:
                        self.logger.warning(
                            "Transfer rule '%s' matched category '%s' but account resolution failed "
                            "(from_number=%s/%s, to_number=%s/%s); importing as expense.",
                            transfer_rule.name,
                            category_name,
                            transfer_rule.from_account_number,
                            transfer_rule.from_account_type,
                            transfer_rule.to_account_number,
                            transfer_rule.to_account_type,
                        )

                txn = TransactionInput(
                    txn_date=txn_date,
                    amount_cents=amount_cents,
                    txn_type="income" if amount_cents > 0 else "expense",
                    payee=internal_payee,
                    description=description or None,
                    category_id=category_id,
                    account_id=account_id,
                    is_subscription=is_subscription,
                    payment_type=payment_type or None,
                    tax_deductible=tax_flag,
                    tax_category=tax_category,
                    note=note_text or None,
                    source_system="xlsx_import",
                    source_uid=source_uid,
                )
                parsed_items.append(("transaction", txn))
                parsed_dates.append(txn_date)
                row += 1

        if not parsed_items:
            raise ValueError("No transaction rows were found in worksheet 'Transactions'.")

        import_period_key = self._infer_import_period_key(parsed_dates)

        deleted_count = 0
        if replace_monthly_baseline:
            # Spreadsheet import is authoritative for the selected sheet period key:
            # remove all prior rows for that period key, regardless of source or txn_date month.
            deleted_count = self.transactions_service.replace_transactions_for_period(
                import_period_key=import_period_key
            )

        imported_count = 0
        for item_kind, posting in parsed_items:
            source_uid = getattr(posting, "source_uid", None)
            prefixed_source_uid = (
                f"xlsx_import:{import_period_key}:{source_uid}" if source_uid else None
            )
            if item_kind == "transfer":
                transfer = replace(
                    posting,
                    import_period_key=import_period_key,
                    source_uid=prefixed_source_uid,
                )
                self.transactions_service.add_transfer(transfer)
                imported_count += 2
            else:
                txn = replace(
                    posting,
                    import_period_key=import_period_key,
                    source_uid=prefixed_source_uid,
                )
                self.transactions_service.add_transaction(txn)
                imported_count += 1

        return XLSXImportResult(
            imported_count=imported_count,
            deleted_count=deleted_count,
            import_period_key=import_period_key,
            year_month_keys=(import_period_key,),
        )
